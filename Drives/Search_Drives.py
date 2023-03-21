import os # to connect/ to communicate with pc
import openpyxl as xl
class SearchDrives():
    '''
    this module helps to find all drives present in pc
    67 to 91 for ascii value
    chr() to convert ascii value to chr
    '''
    def __init__(self):
        self.drives=[] #to store the drives
        self.workbook=xl.load_workbook("C://testdata//Testdrives.xlsx") #to load the worksheet to self.workbook
    def searchmydrives(self)->list:
        for i in range(67,91): #A=65,B=66,C=67,.....Z=90
            if os.path.exists(chr(i)+":"): #'C:'
                self.drives.append(chr(i))
        return self.drives
dr=SearchDrives() #creating object
data=str(dr.searchmydrives()) #list object can be converted to string for storing into excel sheet
print(data)
sheet=dr.workbook.active #workbook may contain many sheets among them the active sheet will be assigned
sheet.cell(row=1,column=1).value=data
sheet.cell(row=2,column=1).value=data
sheet.cell(row=3,column=1).value=data
dr.workbook.save("C://testdata//Testdrives.xlsx")
dr.workbook.close()
# dr.searchmydrives()
# print(dr.drives)